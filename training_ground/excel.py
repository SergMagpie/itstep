from math import floor
from openpyxl import load_workbook

tb = load_workbook('d:/itstep/training_ground/trekking3.xlsx')
sheet_names = tb.sheetnames
book = []
for name in sheet_names:
    dates = []
    for row in tb[name].iter_rows():
        rows = []
        for cell in row:
            rows.append(cell.value)
        dates.append(rows)
    book.append(dates)
kalories = {}
for i in book[0]:
    kalories[i[0]] = i[1:]
a = []
for i in book[1][1:]:
    a.append([i[0]] + [i[2] * j / 100 if j else 0 for j in kalories[i[1]]])
# print(a)
days = []
for i in a:
    days.append(i[0])
days = list(set(days))
for day in days:
    rez = [0, 0, 0, 0]
    for i in a:
        for n, j in enumerate(i[1:]):
            if i[0] == day:
                rez[n] += j
    print(*[floor(i) for i in rez])
